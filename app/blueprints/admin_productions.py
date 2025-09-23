# app/blueprints/admin_productions.py
from flask import (
    Blueprint, render_template, request, session, abort,
    send_file, flash, redirect, url_for
)
from ..repositories.productions import ProductionRepository
from ..services.hospital_service import HospitalService
from ..services.user_service import UserService
from ..services.procedure_service import ProcedureService
from ..repositories.hospital_prices import HospitalPriceRepository
from ..db import get_conn

# --- Excel ---
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from datetime import datetime, date
from typing import Optional

bp = Blueprint("admin_productions", __name__)
repo = ProductionRepository()
hsvc = HospitalService()
usvc = UserService()
psvc = ProcedureService()
price_repo = HospitalPriceRepository()


def _admin_required():
    return bool(session.get("user_id")) and session.get("role") == "admin"


@bp.before_request
def guard():
    if not _admin_required():
        abort(403)


# ---------------------------
# Helpers
# ---------------------------
def _norm_date(s: Optional[str]) -> Optional[str]:
    s = (s or "").strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def _money(s: Optional[str]) -> Optional[str]:
    s = (s or "").strip()
    if not s:
        return None
    # troca vírgula por ponto e remove milhar se vier "1.234,56"
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    return s


def _find_hospital_id(cur, key) -> Optional[int]:
    if key is None:
        return None
    # ID numérico?
    if isinstance(key, int) or (isinstance(key, str) and key.isdigit()):
        cur.execute("SELECT id FROM hospitals WHERE id=%s;", (int(key),))
        row = cur.fetchone()
        return row["id"] if row else None
    k = (str(key) or "").strip().lower()
    if not k:
        return None
    cur.execute("""
       SELECT id
         FROM hospitals
        WHERE LOWER(COALESCE(nickname,'')) = %s
           OR LOWER(COALESCE(trade_name,'')) = %s
           OR LOWER(COALESCE(corporate_name,'')) = %s
        LIMIT 1;
    """, (k, k, k))
    row = cur.fetchone()
    return row["id"] if row else None


def _find_doctor_user_id(cur, key: Optional[str]) -> Optional[int]:
    k = (key or "").strip().lower()
    if not k:
        return None
    # username
    cur.execute("SELECT id FROM users WHERE role='doctor' AND LOWER(username)=%s LIMIT 1;", (k,))
    row = cur.fetchone()
    if row:
        return row["id"]
    # nome completo
    cur.execute("""
        SELECT u.id
          FROM users u
          LEFT JOIN doctors d ON d.user_id=u.id
         WHERE u.role='doctor' AND LOWER(COALESCE(d.full_name,''))=%s
         LIMIT 1;
    """, (k,))
    row = cur.fetchone()
    return row["id"] if row else None


def _find_procedure_id(cur, key: Optional[str]) -> Optional[int]:
    k = (key or "").strip()
    if not k:
        return None
    # TUSS
    cur.execute("SELECT id FROM procedures WHERE tuss_code=%s LIMIT 1;", (k,))
    row = cur.fetchone()
    if row:
        return row["id"]
    # nome
    cur.execute("SELECT id FROM procedures WHERE LOWER(name)=LOWER(%s) LIMIT 1;", (k,))
    row = cur.fetchone()
    return row["id"] if row else None


# ---------------------------
# Listagem
# ---------------------------
@bp.route("/productions", methods=["GET"])
def list_all():
    doctor_id = request.args.get("doctor_id")
    hospital_id = request.args.get("hospital_id")
    date_from = request.args.get("date_from")
    date_to   = request.args.get("date_to")

    did = int(doctor_id) if doctor_id and doctor_id.isdigit() else None
    hid = int(hospital_id) if hospital_id and hospital_id.isdigit() else None

    rows = repo.list(
        doctor_user_id=did,
        hospital_id=hid,
        date_from=date_from or None,
        date_to=date_to or None,
        limit=10000
    )

    hospitals = hsvc.list("")  # todos
    users = usvc.list_users()
    doctors = [u for u in users if u.get("role") == "doctor"]

    return render_template(
        "admin/productions_list.html",
        rows=rows,
        hospitals=hospitals,
        doctors=doctors,
        sel_doctor=did, sel_hospital=hid,
        date_from=date_from or "", date_to=date_to or ""
    )


# ---------------------------
# Exportar Excel
# ---------------------------
@bp.route("/productions/export.xlsx", methods=["GET"])
def export_xlsx():
    doctor_id = request.args.get("doctor_id")
    hospital_id = request.args.get("hospital_id")
    date_from = request.args.get("date_from")
    date_to   = request.args.get("date_to")

    did = int(doctor_id) if doctor_id and doctor_id.isdigit() else None
    hid = int(hospital_id) if hospital_id and hospital_id.isdigit() else None

    rows = repo.list(
        doctor_user_id=did,
        hospital_id=hid,
        date_from=date_from or None,
        date_to=date_to or None,
        limit=100000
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Produção"

    header = [
        "Data", "Hospital", "Médico", "TUSS", "Procedimento",
        "Quantidade", "Vlr Unit. (R$)", "Total (R$)", "Obs."
    ]
    ws.append(header)

    for r in rows:
        ws.append([
            str(r.get("exec_date") or ""),
            r.get("hospital_name") or "",
            r.get("doctor_name") or r.get("username") or "",
            r.get("tuss_code") or "",
            r.get("procedure_name") or "",
            r.get("quantity") or 0,
            float(r.get("unit_price") or 0),
            float(r.get("total") or 0),
            r.get("note") or "",
        ])

    # formatos numéricos
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=8):
        if row[0].value is not None:
            row[0].number_format = "0"
        if row[1].value is not None:
            row[1].number_format = "#,##0.00"
        if row[2].value is not None:
            row[2].number_format = "#,##0.00"

    # larguras
    for idx, col_cells in enumerate(ws.iter_cols(min_col=1, max_col=ws.max_column), start=1):
        max_len = 0
        for cell in col_cells:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(12, max_len + 2), 60)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    if date_from or date_to:
        fname = f"producao_{date_from or 'ini'}_a_{date_to or 'fim'}.xlsx"
    else:
        fname = "producao.xlsx"

    return send_file(
        bio,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ---------------------------
# Download do Modelo (.xlsx)
# ---------------------------
@bp.get("/productions/import/template")
def download_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "import"

    headers = [
        "data",            # YYYY-MM-DD ou DD/MM/AAAA
        "hospital",        # apelido (nickname) OU ID
        "medico",          # username OU nome completo
        "procedimento",    # TUSS (preferido) OU nome exato
        "quantidade",      # inteiro (default 1)
        "valor_unitario",  # opcional; se vazio, tenta tabela de preços
        "obs"              # opcional
    ]
    ws.append(headers)
    ws.append(["2025-01-15", "EyeCenter Botucatu", "drjoao", "0405050380", 1, "120,50", "exemplo"])
    ws.append(["15/02/2025", 1, "maria.silva", "Biometria", 2, "", "sem valor — usa tabela"])

    # largura
    for idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 24

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        as_attachment=True,
        download_name="modelo_importacao_producao.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ---------------------------
# Importar Excel
# ---------------------------
@bp.post("/productions/import")
def import_excel():
    file = request.files.get("file")
    if not file or not file.filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        flash("Envie um arquivo .xlsx válido.", "error")
        return redirect(url_for("admin_productions.list_all"))

    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active
    except Exception as e:
        flash(f"Não foi possível ler o Excel: {e}", "error")
        return redirect(url_for("admin_productions.list_all"))

    header = [str(ws.cell(row=1, column=c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]

    def col_idx(name: str) -> Optional[int]:
        try:
            return header.index(name) + 1
        except ValueError:
            return None

    required = ["data", "hospital", "medico", "procedimento"]
    missing = [h for h in required if h not in header]
    if missing:
        flash(f"Cabeçalho inválido. Faltando colunas: {', '.join(missing)}", "error")
        return redirect(url_for("admin_productions.list_all"))

    c_data = col_idx("data")
    c_hosp = col_idx("hospital")
    c_med  = col_idx("medico")
    c_proc = col_idx("procedimento")
    c_qty  = col_idx("quantidade")
    c_val  = col_idx("valor_unitario")
    c_obs  = col_idx("obs")

    ok_count = 0
    err_rows = []

    with get_conn() as conn, conn.cursor() as cur:
        for r in range(2, ws.max_row + 1):
            def cell(ci):
                return (ws.cell(row=r, column=ci).value if ci else None)

            raw_date = cell(c_data)
            raw_h    = cell(c_hosp)
            raw_m    = cell(c_med)
            raw_p    = cell(c_proc)
            raw_q    = cell(c_qty)
            raw_v    = cell(c_val)
            raw_o    = cell(c_obs)

            # linha vazia?
            if not (raw_date or raw_h or raw_m or raw_p):
                continue

            # data
            if isinstance(raw_date, (datetime, date)):
                exec_date = raw_date.date().isoformat() if isinstance(raw_date, datetime) else raw_date.isoformat()
            else:
                exec_date = _norm_date(str(raw_date))

            qty = int(raw_q) if str(raw_q).strip().isdigit() else 1
            unit_price = _money(str(raw_v)) if raw_v is not None else None
            note = str(raw_o).strip() if raw_o is not None else None

            hid = _find_hospital_id(cur, raw_h)
            did = _find_doctor_user_id(cur, str(raw_m) if raw_m is not None else "")
            pid = _find_procedure_id(cur, str(raw_p) if raw_p is not None else "")

            problems = []
            if not exec_date:
                problems.append("data inválida")
            if not hid:
                problems.append("hospital não encontrado")
            if not did:
                problems.append("médico não encontrado")
            if not pid:
                problems.append("procedimento não encontrado")

            if problems:
                err_rows.append(f"L{r}: " + ", ".join(problems))
                continue

            # tenta tabela de preços se não veio valor
            if not unit_price:
                try:
                    price = price_repo.resolve_price(hid, pid)
                    if price is not None:
                        unit_price = str(price)
                except Exception:
                    pass

            try:
                cur.execute(
                    """
                    INSERT INTO productions
                      (exec_date, doctor_user_id, hospital_id, procedure_id, quantity, unit_price, note)
                    VALUES
                      (%s, %s, %s, %s, %s, NULLIF(%s,'')::numeric, NULLIF(%s,''));
                    """,
                    (exec_date, did, hid, pid, qty, unit_price or "", note or ""),
                )
                ok_count += 1
            except Exception as e:
                err_rows.append(f"L{r}: erro ao inserir ({e})")

    if ok_count:
        flash(f"Importação concluída: {ok_count} linha(s) inserida(s).", "ok")
    if err_rows:
        resumo = "; ".join(err_rows[:10])
        mais = f" (+{len(err_rows)-10}…)" if len(err_rows) > 10 else ""
        flash(f"Linhas ignoradas: {len(err_rows)}. {resumo}{mais}", "error")

    return redirect(url_for("admin_productions.list_all"))
